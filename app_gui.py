import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from datetime import datetime, date
import pathlib


# ============================
#    Funciones auxiliares
# ============================

def evaluate_simple_formula(formula_text):
    if not formula_text:
        return ""
    f = str(formula_text).strip().lstrip('=').upper()

    if 'TODAY' in f or 'HOY' in f:
        return date.today()
    if 'NOW' in f or 'AHORA' in f:
        return datetime.now()

    return ""


def extract_fields_from_docx(doc_path):
    """Obtiene placeholders {{...}} simples desde el XML."""
    from zipfile import ZipFile
    import re

    tags = []
    try:
        with ZipFile(doc_path, 'r') as z:
            for name in z.namelist():
                if name.startswith("word/") and name.endswith(".xml"):
                    xml = z.read(name).decode('utf-8', errors='ignore')
                    for m in re.findall(r"\{\{\s*([^{}]+?)\s*\}\}", xml):
                        tags.append(m.strip())
    except:
        pass
    return list(dict.fromkeys(tags))


def read_excel_mapping(xlsx_path, log):
    """Lee hoja 'mapeo' del Excel y devuelve diccionario etiqueta->celda."""
    wb_vals = load_workbook(xlsx_path, data_only=True)
    wb_forms = load_workbook(xlsx_path, data_only=False)

    # detectar hoja mapeo
    map_sheet = None
    for name in wb_vals.sheetnames:
        if name.lower().strip() == "mapeo":
            map_sheet = wb_vals[name]
            break
    if map_sheet is None:
        map_sheet = wb_vals[wb_vals.sheetnames[0]]

    mapping = {}
    for row in map_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        tag_raw = row[0]
        coord_raw = row[1]
        if tag_raw:
            mapping[str(tag_raw).strip()] = str(coord_raw).strip() if coord_raw else ""

    # hoja medicamento = la primera después de mapeo
    meds = [n for n in wb_vals.sheetnames if n != map_sheet.title]
    default_med_sheet = meds[0] if meds else map_sheet.title

    return mapping, wb_vals, wb_forms, default_med_sheet


def read_cell(wb_vals, wb_forms, sheet, coord):
    """Lee una celda o rango, evaluando fórmulas simples."""
    from openpyxl.utils import range_boundaries, coordinate_to_tuple

    ws_vals = wb_vals[sheet]
    ws_forms = wb_forms[sheet]

    # rango
    if ":" in coord:
        try:
            min_c, min_r, max_c, max_r = range_boundaries(coord)
        except:
            return ""

        rows_out = []
        for r in range(min_r, max_r + 1):
            row_vals = []
            for c in range(min_c, max_c + 1):
                cell = ws_vals.cell(row=r, column=c)
                val = cell.value

                fcell = ws_forms.cell(row=r, column=c)
                if (val is None or val == "") and (fcell.data_type == "f" or str(fcell.value).startswith("=")):
                    val = evaluate_simple_formula(fcell.value)

                row_vals.append(val if val is not None else "")
            rows_out.append("\t".join([str(v) for v in row_vals]))
        return "\n".join(rows_out)
    else:
        # celda simple
        try:
            cell_v = ws_vals[coord].value
        except:
            return ""

        fcell = ws_forms[coord]
        if (cell_v is None or cell_v == "") and (fcell.data_type == "f" or str(fcell.value).startswith("=")):
            cell_v = evaluate_simple_formula(fcell.value)

        return "" if cell_v is None else str(cell_v)


# ====================================
#      PROCESADOR PRINCIPAL
# ====================================

def procesar_archivo(docx_path, xlsx_path, output_dir, log):
    try:
        log(" > Leyendo plantilla...")
        tags = extract_fields_from_docx(docx_path)

        log(" > Analizando Excel...")
        mapping, wb_vals, wb_forms, meds_sheet = read_excel_mapping(xlsx_path, log)

        context = {}

        for tag in tags:
            coord = mapping.get(tag, "")
            if not coord:
                context[tag] = ""
                continue

            # hoja!coord
            if "!" in coord:
                sheet, cell = coord.split("!", 1)
            else:
                sheet = meds_sheet
                cell = coord

            try:
                val = read_cell(wb_vals, wb_forms, sheet, cell)
            except:
                val = ""

            context[tag] = val

        log(" > Rellenando documento...")

        doc = DocxTemplate(docx_path)
        doc.render(context)

        base = os.path.splitext(os.path.basename(xlsx_path))[0]
        out_name = f"{base}_GENERADO.docx"
        out_path = os.path.join(output_dir, out_name)

        doc.save(out_path)

        log(f" ✔ Documento generado: {out_path}")

        return True

    except Exception as e:
        log(f"ERROR FATAL: {e}")
        return False


# ====================================
#        INTERFAZ GRAFICA
# ====================================

class App:
    def __init__(self, root):
        self.root = root
        root.title("Generador Automático de Protocolos v1.0")
        root.geometry("850x600")

        self.docx_path = tk.StringVar()
        self.xlsx_files = []

        # Crear carpeta segura en Documentos
        documents = pathlib.Path.home() / "Documents" / "Protocolos_Generados"
        os.makedirs(documents, exist_ok=True)
        self.output_dir = str(documents)

        # -------------------------
        #      UI
        # -------------------------
        tk.Label(root, text="Generador Automático de Protocolos v1.0",
                 font=("Arial", 18, "bold")).pack(pady=10)

        frame = tk.Frame(root)
        frame.pack(pady=5)

        # PLANTILLA DOCX
        tk.Label(frame, text="Plantilla (.docx):", font=("Arial", 12, "bold")).grid(row=0, column=0, sticky="w")
        tk.Entry(frame, textvariable=self.docx_path, width=70).grid(row=0, column=1, padx=5)
        tk.Button(frame, text="Seleccionar...", command=self.seleccionar_docx).grid(row=0, column=2, padx=5)

        # EXCELS
        tk.Label(frame, text="Archivos Excel (.xlsx):", font=("Arial", 12, "bold")).grid(row=1, column=0, sticky="nw")
        self.listbox = tk.Listbox(frame, width=70, height=5)
        self.listbox.grid(row=1, column=1, padx=5, pady=5)

        tk.Button(frame, text="Agregar Excel...", command=self.agregar_excel).grid(row=1, column=2, padx=5)

        # BOTÓN GENERAR
        tk.Button(root, text="GENERAR DOCUMENTOS", font=("Arial", 14, "bold"),
                  command=self.generar_documentos).pack(pady=10)

        # LOG
        tk.Label(root, text="Registro / Log:", font=("Arial", 12, "bold")).pack()
        self.log_area = ScrolledText(root, width=100, height=15)
        self.log_area.pack(pady=5)

        self.log(f"Carpeta de salida: {self.output_dir}")

    def log(self, msg):
        self.log_area.insert(tk.END, msg + "\n")
        self.log_area.see(tk.END)

    def seleccionar_docx(self):
        path = filedialog.askopenfilename(filetypes=[("Word", "*.docx")])
        if path:
            self.docx_path.set(path)

    def agregar_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.xlsx_files.append(path)
            self.listbox.insert(tk.END, path)

    def generar_documentos(self):
        if not self.docx_path.get():
            messagebox.showerror("Error", "Selecciona una plantilla .docx")
            return

        if not self.xlsx_files:
            messagebox.showerror("Error", "Agrega al menos un archivo Excel")
            return

        self.log("Iniciando generación...\n")

        ok = 0

        for x in self.xlsx_files:
            self.log(f"Procesando: {x}")
            if procesar_archivo(self.docx_path.get(), x, self.output_dir, self.log):
                ok += 1

        messagebox.showinfo("Proceso finalizado",
                            f"Generados correctamente: {ok}/{len(self.xlsx_files)}\n"
                            f"Carpeta: {self.output_dir}")


# Main
if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
