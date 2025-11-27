# app.py
# Generador de Protocolos - GUI Avanzada (Tkinter)
# Integra el motor V3/V4 para rellenar plantillas DOCX desde Excel
# Soporta MERGEFIELD residuales, rangos, formulas, validación y multi-excel.
# Diseñado para ser compilado con PyInstaller en Windows.

import os
import re
import unicodedata
import threading
import time
import traceback
import subprocess
from zipfile import ZipFile
from datetime import datetime, date
from tkinter import Tk, Button, Label, Listbox, END, filedialog, messagebox, ttk, scrolledtext, StringVar
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, coordinate_to_tuple

# ----------------------------
# UTILIDADES
# ----------------------------
def strip_accents(s: str) -> str:
    if not s: return ""
    nfkd = unicodedata.normalize('NFKD', str(s))
    return ''.join([c for c in nfkd if not unicodedata.combining(c)])

def normalize_for_match(s: str) -> str:
    if s is None: return ""
    s2 = strip_accents(str(s)).lower()
    s2 = re.sub(r'[^0-9a-z]+', '_', s2)
    s2 = re.sub(r'_+', '_', s2).strip('_')
    return s2

# ----------------------------
# EXTRAER PLACEHOLDERS Y MERGEFIELDS
# ----------------------------
def extract_visible_placeholders(docx_path):
    placeholders = []
    try:
        with ZipFile(docx_path, 'r') as z:
            for name in z.namelist():
                if name.startswith("word/") and name.endswith(".xml"):
                    xml = z.read(name).decode('utf-8', errors='ignore')
                    for m in re.findall(r"\{\{\s*([^{}]+?)\s*\}\}", xml):
                        placeholders.append(m.strip())
    except Exception as e:
        print("Error extracting visible placeholders:", e)
    return list(dict.fromkeys(placeholders))

def extract_mergefields(docx_path):
    found = []
    try:
        with ZipFile(docx_path, 'r') as z:
            xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
            for m in re.finditer(r'<w:fldSimple[^>]*w:instr="([^"]+)"[^>]*>', xml, flags=re.IGNORECASE):
                instr = m.group(1)
                mm = re.search(r'MERGEFIELD\s+(.+)', instr, flags=re.IGNORECASE)
                if mm:
                    name = mm.group(1).strip().strip('"').strip()
                    name = re.split(r'\\', name)[0].strip()
                    found.append(name)
            for m in re.finditer(r'<w:instrText[^>]*>([^<]+)</w:instrText>', xml, flags=re.IGNORECASE):
                txt = m.group(1)
                mm = re.search(r'MERGEFIELD\s+(.+)', txt, flags=re.IGNORECASE)
                if mm:
                    name = mm.group(1).strip().strip('"').strip()
                    name = re.split(r'\\', name)[0].strip()
                    found.append(name)
            for block in re.finditer(r'(<w:fldChar[^>]*w:fldCharType="begin"[^>]*>.*?</w:fldChar>)(.*?)((?:<w:fldChar[^>]*w:fldCharType="end"[^>]*>.*?</w:fldChar>))', xml, flags=re.IGNORECASE|re.DOTALL):
                seg = block.group(0)
                instrs = re.findall(r'<w:instrText[^>]*>([^<]+)</w:instrText>', seg, flags=re.IGNORECASE)
                if instrs:
                    joined = " ".join(instrs)
                    mm = re.search(r'MERGEFIELD\s+(.+)', joined, flags=re.IGNORECASE)
                    if mm:
                        name = mm.group(1).strip().strip('"').strip()
                        name = re.split(r'\\', name)[0].strip()
                        found.append(name)
    except Exception:
        pass
    cleaned = []
    for f in found:
        if f:
            f2 = f.replace('"', '').strip()
            if f2 not in cleaned:
                cleaned.append(f2)
    return cleaned

# ----------------------------
# LECTURA CELDAS / RANGOS / FORMULAS
# ----------------------------
def evaluate_simple_formula(formula_text):
    if not formula_text: return ""
    f = str(formula_text).strip().lstrip('=').upper()
    if 'TODAY' in f or 'HOY' in f:
        return date.today()
    if 'NOW' in f or 'AHORA' in f:
        return datetime.now()
    return ""

def parse_sheet_and_coord(coord_text, default_sheet):
    if not coord_text or str(coord_text).strip() == '':
        return (default_sheet, '')
    txt = str(coord_text).strip()
    if '!' in txt:
        parts = txt.split('!', 1)
        return (parts[0].strip(), parts[1].strip())
    return (default_sheet, txt)

def read_cell_or_range(wb_vals, wb_forms, sheet_name, coord):
    if sheet_name not in wb_vals.sheetnames:
        return (None, f"Hoja '{sheet_name}' no encontrada")
    ws_vals = wb_vals[sheet_name]
    ws_forms = wb_forms[sheet_name] if sheet_name in wb_forms.sheetnames else None
    coord = str(coord).strip()
    if ':' in coord:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(coord)
        except Exception as e:
            return (None, f"Rango inválido '{coord}': {e}")
        rows_out = []
        for r in range(min_row, max_row+1):
            cols = []
            for c in range(min_col, max_col+1):
                try:
                    cell = ws_vals.cell(row=r, column=c)
                    v = cell.value
                    formula_text = None
                    if ws_forms:
                        try:
                            fcell = ws_forms.cell(row=r, column=c)
                            if fcell.data_type == 'f' or (isinstance(fcell.value, str) and str(fcell.value).startswith('=')):
                                formula_text = fcell.value
                        except:
                            formula_text = None
                    if (v is None or (isinstance(v, str) and str(v).strip() == '')) and formula_text:
                        v = evaluate_simple_formula(formula_text)
                    if isinstance(v, (datetime, date)):
                        v = v.strftime("%d/%m/%Y %H:%M:%S") if isinstance(v, datetime) else v.strftime("%d/%m/%Y")
                except Exception:
                    v = ""
                cols.append("" if v is None else v)
            rows_out.append("\t".join([str(x) for x in cols]))
        return ("\n".join(rows_out), None)
    else:
        try:
            cell = ws_vals[coord]
            v = cell.value
        except Exception:
            try:
                r, cidx = coordinate_to_tuple(coord)
                v = ws_vals.cell(row=r, column=cidx).value
            except Exception as e:
                return (None, f"Coordenada inválida '{coord}': {e}")
        formula_text = None
        if ws_forms:
            try:
                fcell = ws_forms[coord]
                if fcell.data_type == 'f' or (isinstance(fcell.value, str) and str(fcell.value).startswith('=')):
                    formula_text = fcell.value
            except:
                formula_text = None
        if (v is None or (isinstance(v, str) and str(v).strip() == '')) and formula_text:
            v = evaluate_simple_formula(formula_text)
        if isinstance(v, (datetime, date)):
            v = v.strftime("%d/%m/%Y %H:%M:%S") if isinstance(v, datetime) else v.strftime("%d/%m/%Y")
        return ("" if v is None else v, None)

# ----------------------------
# PROCESAR 1 EXCEL -> DEVUELVE CONTEXTO Y REPORTE
# ----------------------------
def procesar_un_excel(xlsx_path, docx_path):
    try:
        visible_tags = extract_visible_placeholders(docx_path)
        mergefields = extract_mergefields(docx_path)
        synthesized_from_merge = [re.sub(r'\s+', '_', mf.strip()) for mf in mergefields]
        template_tags = list(dict.fromkeys(visible_tags + synthesized_from_merge + mergefields))

        wb_values = load_workbook(xlsx_path, data_only=True)
        wb_forms = load_workbook(xlsx_path, data_only=False)

        # determinar hoja mapeo
        mapeo_sheet = None
        for name in wb_values.sheetnames:
            if name.strip().lower() == 'mapeo':
                mapeo_sheet = wb_values[name]
                break
        if mapeo_sheet is None:
            for name in wb_values.sheetnames:
                if 'mapeo' in name.lower():
                    mapeo_sheet = wb_values[name]
                    break
        if mapeo_sheet is None:
            mapeo_sheet = wb_values[wb_values.sheetnames[0]]

        other_sheets = [n for n in wb_values.sheetnames if wb_values[n] != mapeo_sheet]
        default_med_sheet = other_sheets[0] if other_sheets else mapeo_sheet.title

        mapping = []
        for row in mapeo_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            etiqueta_raw = row[0]
            coord_raw = row[1]
            if etiqueta_raw is None:
                continue
            etiqueta = str(etiqueta_raw).strip()
            coord = str(coord_raw).strip() if coord_raw is not None else ""
            mapping.append((etiqueta, coord))

        if not mapping:
            raise RuntimeError("La hoja 'mapeo' no contiene pares etiqueta/coord (col A, col B).")

        excel_map = {tag: coord for tag, coord in mapping}
        norm_excel_map = {normalize_for_match(k): (k, excel_map[k]) for k in excel_map}

        template_to_excel = {}
        report_matches = []
        norm_excel_keys = list(norm_excel_map.keys())

        for t in template_tags:
            t_original = t
            t_norm = normalize_for_match(t_original)
            chosen = None
            if t_norm in norm_excel_map:
                chosen = norm_excel_map[t_norm][0]
                report_matches.append((t_original, chosen, 'norm_exact'))
            else:
                alt = t_original.replace(' ', '_')
                if normalize_for_match(alt) in norm_excel_map:
                    chosen = norm_excel_map[normalize_for_match(alt)][0]
                    report_matches.append((t_original, chosen, 'space_to_underscore'))
                else:
                    alt2 = t_original.replace('_', ' ')
                    if normalize_for_match(alt2) in norm_excel_map:
                        chosen = norm_excel_map[normalize_for_match(alt2)][0]
                        report_matches.append((t_original, chosen, 'underscore_to_space'))
                    else:
                        t_compact = t_norm.replace('_', '')
                        for ek in norm_excel_keys:
                            if ek.replace('_', '') == t_compact:
                                chosen = norm_excel_map[ek][0]
                                report_matches.append((t_original, chosen, 'compact_numeric'))
                                break
            template_to_excel[t_original] = chosen

        context = {}
        errors = []
        empty_cells = []
        used_excel_tags = set()
        for t_orig, excel_tag in template_to_excel.items():
            if excel_tag is None:
                context[t_orig] = ""
                continue
            coord_text = excel_map.get(excel_tag, "")
            sheet_name, coord = parse_sheet_and_coord(coord_text, default_med_sheet)
            if coord == "" or coord is None:
                context[t_orig] = ""
                empty_cells.append((t_orig, f"{sheet_name}!{coord}", "coordenada vacía"))
                used_excel_tags.add(excel_tag)
                continue
            val, err = read_cell_or_range(wb_values, wb_forms, sheet_name, coord)
            if err is not None:
                errors.append((t_orig, excel_tag, sheet_name, coord, err))
                context[t_orig] = ""
            else:
                context[t_orig] = "" if val is None else ("\n".join(val) if isinstance(val, (list, tuple)) else str(val))
                if context[t_orig].strip() == "":
                    empty_cells.append((t_orig, f"{sheet_name}!{coord}", "celda vacía"))
                used_excel_tags.add(excel_tag)

        template_without_mapping = [t for t, v in template_to_excel.items() if v is None]
        excel_not_used = [k for k in excel_map.keys() if k not in used_excel_tags]

        report_lines = []
        report_lines.append(f"Plantilla: {docx_path}")
        report_lines.append(f"Excel: {xlsx_path}")
        report_lines.append(f"Hoja mapeo: {mapeo_sheet.title}")
        report_lines.append(f"Hoja medicamento por defecto: {default_med_sheet}")
        report_lines.append("")
        report_lines.append("Emparejamientos detectados (plantilla -> excel):")
        for a, b, c in report_matches:
            report_lines.append(f" - {a}  =>  {b}   (metodo: {c})")
        report_lines.append("")
        report_lines.append("Plantilla sin mapeo tras heurísticas:")
        for t in template_without_mapping:
            report_lines.append(f" - {t}")
        report_lines.append("")
        report_lines.append("Etiquetas en mapeo Excel no usadas:")
        for t in excel_not_used[:200]:
            report_lines.append(f" - {t}")
        report_lines.append("")
        report_lines.append(f"Celdas vacías: {len(empty_cells)}")
        for it in empty_cells[:200]:
            report_lines.append(f" - etiqueta plantilla: {it[0]} coord: {it[1]} -> {it[2]}")
        report_lines.append("")
        report_lines.append(f"Errores leyendo coordenadas: {len(errors)}")
        for e in errors[:200]:
            report_lines.append(f" - tpl: {e[0]} excel_tag: {e[1]} hoja: {e[2]} coord: {e[3]} -> {e[4]}")

        report_text = "\n".join(report_lines)

        return {
            "context": context,
            "report_text": report_text,
            "report_lines": report_lines,
            "template_tags": template_tags,
            "template_to_excel": template_to_excel,
            "excel_map": excel_map,
            "default_med_sheet": default_med_sheet
        }
    except Exception as e:
        return {"error": str(e), "trace": traceback.format_exc()}

# ----------------------------
# GUI (Tkinter) - INTERFAZ AVANZADA
# ----------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title("Generador de Protocolos - Portable")
        root.geometry("900x700")

        # Variables
        self.template_path = ""
        self.excel_paths = []
        self.output_dir = os.path.join(os.getcwd(), "salidas")
        os.makedirs(self.output_dir, exist_ok=True)

        # Widgets
        Label(root, text="Plantilla (DOCX):").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        Button(root, text="Seleccionar plantilla", command=self.select_template).grid(row=0, column=1, sticky="w")

        Label(root, text="Excels (múltiples):").grid(row=1, column=0, sticky="w", padx=8)
        Button(root, text="Seleccionar Excels", command=self.select_excels).grid(row=1, column=1, sticky="w")

        self.lbl_template = Label(root, text="Ninguna plantilla seleccionada", fg="gray")
        self.lbl_template.grid(row=0, column=2, columnspan=4, sticky="w")

        self.lst_excels = Listbox(root, width=60, height=5)
        self.lst_excels.grid(row=2, column=0, columnspan=6, padx=8, pady=6, sticky="w")

        self.progress = ttk.Progressbar(root, orient='horizontal', length=520, mode='determinate')
        self.progress.grid(row=3, column=0, columnspan=4, padx=8, pady=6, sticky="w")

        self.btn_process = Button(root, text="PROCESAR TODO", bg="#2b8a3e", fg="white", command=self.start_processing)
        self.btn_process.grid(row=3, column=4, padx=8)

        Label(root, text="Reporte:").grid(row=4, column=0, sticky="w", padx=8, pady=6)
        self.txt_report = scrolledtext.ScrolledText(root, width=110, height=20)
        self.txt_report.grid(row=5, column=0, columnspan=6, padx=8, pady=6)

        Label(root, text="Archivos generados:").grid(row=6, column=0, sticky="w", padx=8)
        self.lst_out = Listbox(root, width=80, height=6)
        self.lst_out.grid(row=7, column=0, columnspan=4, padx=8, pady=6, sticky="w")
        Button(root, text="Abrir carpeta de salida", command=self.open_output_folder).grid(row=7, column=4, sticky="w")

    def select_template(self):
        p = filedialog.askopenfilename(filetypes=[("Word templates","*.docx")])
        if p:
            self.template_path = p
            self.lbl_template.config(text=os.path.basename(p), fg="black")

    def select_excels(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel files","*.xlsx;*.xls")])
        if files:
            self.excel_paths = list(files)
            self.lst_excels.delete(0, END)
            for f in self.excel_paths:
                self.lst_excels.insert(END, os.path.basename(f))

    def open_output_folder(self):
        path = self.output_dir
        try:
            if os.name == 'nt':
                os.startfile(path)
            elif os.name == 'posix':
                subprocess.call(['xdg-open', path])
            else:
                messagebox.showinfo("Abrir carpeta", f"Abre manualmente: {path}")
        except Exception:
            messagebox.showinfo("Abrir carpeta", f"Abre manualmente: {path}")

    def start_processing(self):
        if not self.template_path:
            messagebox.showwarning("Falta plantilla", "Selecciona una plantilla .docx primero.")
            return
        if not self.excel_paths:
            messagebox.showwarning("Falta Excel", "Selecciona al menos un archivo Excel.")
            return
        # run in thread to avoid freezing UI
        thread = threading.Thread(target=self.process_all, daemon=True)
        thread.start()

    def process_all(self):
        self.btn_process.config(state="disabled")
        total = len(self.excel_paths)
        self.progress['maximum'] = total
        self.progress['value'] = 0
        self.txt_report.delete(1.0, END)
        self.lst_out.delete(0, END)

        for idx, xlsx in enumerate(self.excel_paths, start=1):
            try:
                base = os.path.splitext(os.path.basename(xlsx))[0]
                self.append_report(f"\n--- Procesando {base} ({idx}/{total}) ---\n")
                result = procesar_un_excel(xlsx, self.template_path)
                if 'error' in result:
                    self.append_report(f"ERROR procesando {base}: {result['error']}\n{result.get('trace','')}\n")
                    continue
                context = result['context']
                report_text = result['report_text']

                # render
                doc = DocxTemplate(self.template_path)
                ctx_for_docxtpl = {k: ("" if v is None else str(v)) for k, v in context.items()}
                try:
                    doc.render(ctx_for_docxtpl)
                except Exception as e:
                    self.append_report(f"Warning: error renderizando plantilla para {base}: {e}\n")

                out_docx = os.path.join(self.output_dir, f"{base}_RELLENADO.docx")
                doc.save(out_docx)

                report_file = os.path.join(self.output_dir, f"{base}_REPORTE.txt")
                with open(report_file, 'w', encoding='utf-8') as f:
                    f.write(report_text)

                self.lst_out.insert(END, os.path.basename(out_docx))
                self.append_report(f"Generado: {out_docx}\nReporte: {report_file}\n")
                # attempt to open the files? we only list them.

                # Try to download automatically if running in environment that supports it (only in web notebooks)
                self.try_download_detection(out_docx, report_file)

            except Exception as e:
                self.append_report(f"Error inesperado en {xlsx}: {str(e)}\n{traceback.format_exc()}\n")
            finally:
                self.progress['value'] = idx
                time.sleep(0.2)

        self.append_report("\n--- PROCESO FINALIZADO ---\n")
        messagebox.showinfo("Listo", "Procesamiento finalizado.")
        self.btn_process.config(state="normal")

    def try_download_detection(self, out_docx, report_file):
        # If running inside Colab/Notebook, files.download would be available;
        # here we just keep outputs local for EXE usage. We still show paths.
        pass

    def append_report(self, text):
        self.txt_report.insert(END, text)
        self.txt_report.see(END)

def main():
    root = Tk()
    app = App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
