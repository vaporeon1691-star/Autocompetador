import os
import re
import unicodedata
from zipfile import ZipFile
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, coordinate_to_tuple
from docxtpl import DocxTemplate


# ============================================================
# Normalización
# ============================================================

def strip_accents(s):
    if not s:
        return ""
    nfkd = unicodedata.normalize('NFKD', s)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])


def normalize_for_match(s):
    if s is None:
        return ""
    s = strip_accents(str(s)).lower()
    s = re.sub(r"[^0-9a-z]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


# ============================================================
# Extraer placeholders visibles {{...}}
# ============================================================

def extract_visible_placeholders(docx_path):
    placeholders = []
    try:
        with ZipFile(docx_path, 'r') as z:
            for name in z.namelist():
                if name.startswith("word/") and name.endswith(".xml"):
                    xml = z.read(name).decode("utf-8", errors="ignore")
                    for m in re.findall(r"\{\{\s*([^{}]+?)\s*\}\}", xml):
                        placeholders.append(m.strip())
    except:
        pass
    return list(dict.fromkeys(placeholders))


# ============================================================
# Extraer MERGEFIELD residuales
# ============================================================

def extract_mergefields(docx_path):
    found = []
    try:
        with ZipFile(docx_path, 'r') as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="ignore")

            # fldSimple
            for m in re.finditer(r'<w:fldSimple[^>]*w:instr="([^"]+)"', xml, flags=re.I):
                instr = m.group(1)
                mm = re.search(r"MERGEFIELD\s+(.+)", instr, flags=re.I)
                if mm:
                    name = re.split(r"\\", mm.group(1).strip())[0].strip()
                    found.append(name)

            # instrText
            for m in re.finditer(r"<w:instrText[^>]*>([^<]+)</w:instrText>", xml, flags=re.I):
                instr = m.group(1)
                mm = re.search(r"MERGEFIELD\s+(.+)", instr, flags=re.I)
                if mm:
                    name = re.split(r"\\", mm.group(1).strip())[0].strip()
                    found.append(name)

    except:
        pass

    # limpiar
    cleaned = []
    for f in found:
        if f and f not in cleaned:
            cleaned.append(f)

    return cleaned


# ============================================================
# Fórmulas: hoy(), now(), etc.
# ============================================================

def evaluate_simple_formula(formula_text):
    if not formula_text:
        return ""
    f = str(formula_text).strip().lstrip("=").upper()
    if "TODAY" in f or "HOY" in f:
        return date.today()
    if "NOW" in f or "AHORA" in f:
        return datetime.now()
    return ""


# ============================================================
# Leer celda o rango
# ============================================================

def parse_sheet_and_coord(coord_text, default_sheet):
    if not coord_text or coord_text.strip() == "":
        return default_sheet, ""
    if "!" in coord_text:
        sh, c = coord_text.split("!", 1)
        return sh.strip(), c.strip()
    return default_sheet, coord_text.strip()


def read_cell_or_range(wb_values, wb_forms, sheet, coord):
    if sheet not in wb_values.sheetnames:
        return None, f"Hoja '{sheet}' no encontrada"

    ws_vals = wb_values[sheet]
    ws_forms = wb_forms[sheet] if sheet in wb_forms.sheetnames else None

    # Rango
    if ":" in coord:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(coord)
        except Exception as e:
            return None, f"Rango inválido {coord}: {e}"

        out = []
        for r in range(min_row, max_row + 1):
            row = []
            for c in range(min_col, max_col + 1):
                v = ws_vals.cell(row=r, column=c).value
                formula_text = None
                if ws_forms:
                    fcell = ws_forms.cell(row=r, column=c)
                    if fcell.data_type == 'f' or (isinstance(fcell.value, str) and fcell.value.startswith('=')):
                        formula_text = fcell.value

                if (v is None or v == "") and formula_text:
                    v = evaluate_simple_formula(formula_text)

                if isinstance(v, (datetime, date)):
                    v = v.strftime("%d/%m/%Y")

                row.append("" if v is None else str(v))
            out.append("\t".join(row))

        return "\n".join(out), None

    # Celda individual
    try:
        v = ws_vals[coord].value
    except:
        return None, f"Coordenada inválida '{coord}'"

    formula_text = None
    if ws_forms:
        fcell = ws_forms[coord]
        if fcell.data_type == 'f' or (isinstance(fcell.value, str) and fcell.value.startswith('=')):
            formula_text = fcell.value

    if (v is None or v == "") and formula_text:
        v = evaluate_simple_formula(formula_text)

    if isinstance(v, (datetime, date)):
        v = v.strftime("%d/%m/%Y")

    return "" if v is None else str(v), None


# ============================================================
# Proceso principal POR EXCEL
# ============================================================

def procesar_archivo_excel(template_path, excel_path, output_folder, log_func):
    """
    Procesa 1 archivo Excel y genera:
    - DOCX rellenado
    - Reporte TXT
    """

    log = log_func
    log(f"\nProcesando archivo: {os.path.basename(excel_path)}\n")

    # 1) Cargar Excel
    wb_values = load_workbook(excel_path, data_only=True)
    wb_forms = load_workbook(excel_path, data_only=False)

    # 2) Detectar hoja mapeo
    mapeo_sheet = None
    for name in wb_values.sheetnames:
        if "mapeo" in name.lower():
            mapeo_sheet = wb_values[name]
            break

    if mapeo_sheet is None:
        raise Exception("No se encontró hoja 'mapeo' en el Excel.")

    # 3) Hoja medicamento (primera distinta de mapeo)
    other_sheets = [s for s in wb_values.sheetnames if s != mapeo_sheet.title]
    default_med_sheet = other_sheets[0] if other_sheets else mapeo_sheet.title

    # 4) Leer mapeo
    mapping = []
    for row in mapeo_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        etiqueta, coord = row
        if etiqueta:
            mapping.append((str(etiqueta).strip(), str(coord).strip() if coord else ""))

    excel_map = {tag: coord for tag, coord in mapping}

    # 5) Extraer tags del Word
    visible_tags = extract_visible_placeholders(template_path)
    merge_tags = extract_mergefields(template_path)

    synthesized = [re.sub(r"\s+", "_", m) for m in merge_tags]

    template_tags = list(dict.fromkeys(visible_tags + synthesized + merge_tags))

    # Mapas normalizados
    norm_excel_map = {normalize_for_match(k): (k, excel_map[k]) for k in excel_map.keys()}

    template_to_excel = {}

    for t in template_tags:
        t_norm = normalize_for_match(t)
        if t_norm in norm_excel_map:
            template_to_excel[t] = norm_excel_map[t_norm][0]
        else:
            template_to_excel[t] = None

    # 6) Construcción del contexto
    context = {}
    for t, excel_tag in template_to_excel.items():
        if excel_tag is None:
            context[t] = ""
            continue

        coord_text = excel_map.get(excel_tag, "")
        sheet, coord = parse_sheet_and_coord(coord_text, default_med_sheet)

        val, err = read_cell_or_range(wb_values, wb_forms, sheet, coord)
        if err:
            context[t] = ""
        else:
            context[t] = val

    # 7) Render Word
    doc = DocxTemplate(template_path)
    doc.render({k: str(v) for k, v in context.items()})

    out_name = os.path.splitext(os.path.basename(excel_path))[0] + "_protocolo.docx"
    out_path = os.path.join(output_folder, out_name)
    doc.save(out_path)

    log(f"Documento generado: {out_path}\n")

    return out_path


# ============================================================
# Procesar VARIOS archivos
# ============================================================

def procesar_archivos(template_path, excels_paths, output_folder, log_func):
    resultados = []
    for excel in excels_paths:
        try:
            result = procesar_archivo_excel(template_path, excel, output_folder, log_func)
            resultados.append(result)
        except Exception as e:
            log_func(f"ERROR procesando {excel}:\n{e}\n")

    return resultados
